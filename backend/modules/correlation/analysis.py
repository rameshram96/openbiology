"""
Correlation Module — Python Analysis Engine
Replaces: correlation_analysis.R  and  heatmap_only.R

Outputs (identical filenames to the R version so the frontend never changes):
  correlation_heatmap.png
  scatter_matrix.png
  correlation_results.xlsx
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.colors import LinearSegmentedColormap
import seaborn as sns
from scipy import stats
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from itertools import combinations

# ─── Colorblind-safe palettes (mirrors the R version exactly) ─────────────────

PALETTES = {
    "RdBu": ["#d73027", "#f46d43", "#fdae61", "#ffffff", "#74add1", "#4575b4"],
    "PRGn": ["#762a83", "#af8dc3", "#f7f7f7", "#7fbf7b", "#1b7837"],
    "PiYG": ["#c51b7d", "#e9a3c9", "#f7f7f7", "#a1d76a", "#4d9221"],
    "Colorblind": ["#0072B2", "#56B4E9", "#ffffff", "#E69F00", "#D55E00"],
    "Greyscale": ["#000000", "#888888", "#ffffff"],
    "Blues": ["#f7fbff", "#6baed6", "#08306b"],
}

SCATTER_POINT_COLOR = "#0072B2"
SCATTER_ALPHA       = 0.35
DIAG_FILL_COLOR     = "#009E73"
POS_COR_COLOR       = "#0072B2"
NEG_COR_COLOR       = "#D55E00"


def _make_cmap(palette_name: str):
    colors = PALETTES.get(palette_name, PALETTES["RdBu"])
    return LinearSegmentedColormap.from_list(palette_name, colors, N=256)


def _sig_stars(p: float) -> str:
    if np.isnan(p):
        return ""
    if p < 0.001:
        return "***"
    if p < 0.01:
        return "**"
    if p < 0.05:
        return "*"
    return ""


def _base_ax(ax, show_grid=True):
    ax.set_facecolor("#FAFAFA")
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#999999", labelsize=8)
    if show_grid:
        ax.grid(True, linestyle="--", linewidth=0.4,
                color="#DDDDDD", alpha=0.8, zorder=0)


# ─── Core computation ─────────────────────────────────────────────────────────

def compute_matrices(df: pd.DataFrame, method: str):
    """
    Returns cor_matrix (np), p_matrix (np), sig_matrix (str), variable names.
    method: 'pearson' | 'spearman' | 'kendall'
    """
    cols = list(df.columns)
    p    = len(cols)
    cor  = np.ones((p, p))
    pmat = np.ones((p, p))

    for i, j in combinations(range(p), 2):
        x = df.iloc[:, i].dropna()
        y = df.iloc[:, j].dropna()
        # align on common index
        common = x.index.intersection(y.index)
        xi, yi = x[common].values, y[common].values

        if method == "pearson":
            r, pv = stats.pearsonr(xi, yi)
        elif method == "spearman":
            r, pv = stats.spearmanr(xi, yi)
        else:  # kendall
            r, pv = stats.kendalltau(xi, yi)

        cor[i, j] = cor[j, i] = float(r)
        pmat[i, j] = pmat[j, i] = float(pv)

    sig = np.empty((p, p), dtype=object)
    for i in range(p):
        for j in range(p):
            sig[i, j] = "-" if i == j else _sig_stars(pmat[i, j])

    return cor, pmat, sig, cols


# ─── Heatmap ──────────────────────────────────────────────────────────────────

def _draw_heatmap(
    cor_mat: np.ndarray,
    p_mat: np.ndarray,
    variables: list,
    method: str,
    n: int,
    params: dict,
    out_path: Path,
):
    palette       = params.get("heatmap_palette", "RdBu")
    axis_fs       = float(params.get("axis_font_size", 0.85))
    show_coef     = bool(params.get("show_coef", True))
    plot_title    = params.get("plot_title", "")

    title = plot_title if plot_title else f"{method.upper()} Correlation Heatmap"
    cmap  = _make_cmap(palette)
    p_cnt = len(variables)

    fig_size = max(7, p_cnt * 0.7 + 2)
    fig, axes = plt.subplots(
        2, 1,
        figsize=(fig_size, fig_size * 0.92),
        dpi=300,
        gridspec_kw={"height_ratios": [0.92, 0.08]},
    )
    fig.patch.set_facecolor("#FAFAFA")
    ax, ax_foot = axes

    # Draw upper-triangle heatmap via seaborn
    mask = np.tril(np.ones_like(cor_mat, dtype=bool), k=-1)  # mask lower+diag
    # seaborn needs the full matrix; we mask lower triangle
    mask_lower = np.tril(np.ones_like(cor_mat, dtype=bool))

    im = ax.imshow(
        cor_mat,
        cmap=cmap,
        vmin=-1, vmax=1,
        aspect="auto",
        interpolation="nearest",
    )

    # White-out lower triangle (keep diagonal white too for clean look)
    for i in range(p_cnt):
        for j in range(p_cnt):
            if i > j:
                ax.add_patch(plt.Rectangle(
                    (j - 0.5, i - 0.5), 1, 1,
                    fc="#FAFAFA", ec="#FAFAFA", zorder=2
                ))

    # Cell annotations — coefficient values
    if show_coef:
        for i in range(p_cnt):
            for j in range(i, p_cnt):
                val = cor_mat[i, j]
                txt_color = "white" if abs(val) > 0.55 else "#333333"
                ax.text(j, i, f"{val:.2f}",
                        ha="center", va="center",
                        fontsize=axis_fs * 7.5,
                        color=txt_color, fontweight="bold", zorder=3)

    # Significance stars above the coefficient
    for i in range(p_cnt):
        for j in range(i + 1, p_cnt):
            stars = _sig_stars(p_mat[i, j])
            if stars:
                ax.text(j, i - 0.28, stars,
                        ha="center", va="center",
                        fontsize=axis_fs * 7,
                        color="#1C1C1C", fontweight="bold", zorder=4)

    # Axes labels
    ax.set_xticks(range(p_cnt))
    ax.set_yticks(range(p_cnt))
    ax.set_xticklabels(variables, rotation=45, ha="right",
                       fontsize=axis_fs * 9, color="#333333")
    ax.set_yticklabels(variables, fontsize=axis_fs * 9, color="#333333")
    ax.tick_params(length=0)
    for spine in ax.spines.values():
        spine.set_visible(False)

    # Colorbar
    cbar = fig.colorbar(im, ax=ax, shrink=0.7, pad=0.02, aspect=20)
    cbar.set_label("Correlation", fontsize=8, color="#555555", labelpad=6)
    cbar.ax.tick_params(labelsize=7, colors="#999999")
    cbar.outline.set_edgecolor("#DDDDDD")

    if title:
        ax.set_title(title, fontsize=axis_fs * 11.5,
                     fontweight="bold", color="#1C1C1C", pad=12)

    # Footnote strip
    ax_foot.axis("off")
    ax_foot.text(
        0.5, 0.5,
        f"Significance: * p<0.05   ** p<0.01   *** p<0.001"
        f"   ({method.upper()} correlation, n={n})",
        ha="center", va="center",
        fontsize=axis_fs * 7.8, color="#555555",
        transform=ax_foot.transAxes,
    )

    fig.tight_layout(pad=1.5)
    fig.savefig(out_path, dpi=300, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)


# ─── Scatter matrix ───────────────────────────────────────────────────────────

def _draw_scatter(
    df: pd.DataFrame,
    cor_mat: np.ndarray,
    p_mat: np.ndarray,
    variables: list,
    method: str,
    params: dict,
    out_path: Path,
):
    axis_fs   = float(params.get("axis_font_size", 0.85))
    plot_title= params.get("plot_title", "")
    dpi       = int(params.get("scatter_dpi", 300))
    title     = plot_title if plot_title else f"{method.upper()} Correlation — Scatter Matrix"

    p_cnt     = len(variables)
    cell_size = max(2.0, 8.0 / p_cnt)
    fig_size  = cell_size * p_cnt

    fig, axes = plt.subplots(
        p_cnt, p_cnt,
        figsize=(fig_size, fig_size),
        dpi=dpi,
    )
    fig.patch.set_facecolor("#FAFAFA")

    if p_cnt == 1:
        axes = np.array([[axes]])

    for i in range(p_cnt):
        for j in range(p_cnt):
            ax = axes[i, j]
            ax.set_facecolor("#FAFAFA")
            for sp in ax.spines.values():
                sp.set_color("#DDDDDD")
                sp.set_linewidth(0.5)
            ax.tick_params(labelsize=axis_fs * 6.5, colors="#999999",
                           length=2, width=0.5)

            xi = df.iloc[:, j].values
            yi = df.iloc[:, i].values

            if i == j:
                # Diagonal — density plot
                valid = xi[~np.isnan(xi)]
                if len(valid) > 2:
                    try:
                        from scipy.stats import gaussian_kde
                        kde = gaussian_kde(valid)
                        xr  = np.linspace(valid.min(), valid.max(), 200)
                        ax.fill_between(xr, kde(xr), alpha=0.45,
                                        color=DIAG_FILL_COLOR)
                        ax.plot(xr, kde(xr), color="#333333",
                                linewidth=0.8)
                    except Exception:
                        ax.hist(valid, bins=8, color=DIAG_FILL_COLOR,
                                alpha=0.6, edgecolor="white", linewidth=0.4)
                ax.set_ylabel("") ; ax.set_xlabel("")

            elif i < j:
                # Upper triangle — correlation coefficient + stars
                r   = cor_mat[i, j]
                pv  = p_mat[i, j]
                sig = _sig_stars(pv)
                col = POS_COR_COLOR if r >= 0 else NEG_COR_COLOR
                fs  = axis_fs * 8 + abs(r) * 4   # bigger font for stronger r
                ax.text(0.5, 0.5, f"{r:.3f}{sig}",
                        ha="center", va="center",
                        fontsize=fs, color=col,
                        fontweight="bold",
                        transform=ax.transAxes)
                ax.set_xticks([]); ax.set_yticks([])

            else:
                # Lower triangle — scatter + regression line
                mask = ~(np.isnan(xi) | np.isnan(yi))
                xs, ys = xi[mask], yi[mask]
                ax.scatter(xs, ys,
                           c=SCATTER_POINT_COLOR,
                           s=max(4, 40 / p_cnt),
                           alpha=SCATTER_ALPHA,
                           edgecolors="white",
                           linewidths=0.3,
                           zorder=3)
                if len(xs) > 2:
                    try:
                        sl, ic, *_ = stats.linregress(xs, ys)
                        xr = np.linspace(xs.min(), xs.max(), 200)
                        ax.plot(xr, sl * xr + ic,
                                color="#56B4E9", linewidth=0.9, zorder=4)
                        # Light CI band
                        ax.fill_between(
                            xr,
                            sl * xr + ic - np.std(ys - (sl * xs + ic)),
                            sl * xr + ic + np.std(ys - (sl * xs + ic)),
                            color="#56B4E9", alpha=0.12, zorder=2,
                        )
                    except Exception:
                        pass

            # Axis labels on edges only
            if j == 0:
                ax.set_ylabel(variables[i],
                              fontsize=axis_fs * 7.5,
                              color="#333333", labelpad=3)
            if i == p_cnt - 1:
                ax.set_xlabel(variables[j],
                              fontsize=axis_fs * 7.5,
                              color="#333333", labelpad=3)

    if title:
        fig.suptitle(title, fontsize=axis_fs * 11,
                     fontweight="bold", color="#1C1C1C", y=1.01)

    fig.tight_layout(pad=0.5)
    fig.savefig(out_path, dpi=dpi, bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)


# ─── Excel export ─────────────────────────────────────────────────────────────

def _save_excel(
    cor_mat: np.ndarray,
    p_mat: np.ndarray,
    sig_mat: np.ndarray,
    variables: list,
    method: str,
    n: int,
    out_path: Path,
):
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", fgColor="1C3557")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(bottom=Side(style="thin", color="888888"))
    center      = Alignment(horizontal="center")

    def write_matrix_sheet(ws, title, matrix, fmt):
        # Header row
        ws.append(["Variable"] + variables)
        for cell in ws[1]:
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = thin_border
        # Data rows
        for i, var in enumerate(variables):
            row = [var]
            for j in range(len(variables)):
                v = matrix[i, j]
                if isinstance(v, (float, np.floating)):
                    row.append(round(float(v), 6) if not np.isnan(v) else "—")
                else:
                    row.append(str(v))
            ws.append(row)
        # Column widths
        ws.column_dimensions["A"].width = 22
        for ci, _ in enumerate(variables, start=2):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(ci)
            ].width = 14

    # Sheet 1 — Correlation Matrix
    ws1 = wb.active
    ws1.title = "Correlation Matrix"
    write_matrix_sheet(ws1, "Correlation Matrix", cor_mat, ".4f")

    # Sheet 2 — P-values
    ws2 = wb.create_sheet("P-values")
    write_matrix_sheet(ws2, "P-values", p_mat, ".6f")

    # Sheet 3 — Significance
    ws3 = wb.create_sheet("Significance")
    write_matrix_sheet(ws3, "Significance", sig_mat, "s")

    # Sheet 4 — Summary
    ws4 = wb.create_sheet("Summary")
    ws4.append(["Parameter", "Value"])
    for cell in ws4[1]:
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = thin_border
    for row in [
        ("Method",       method.upper()),
        ("Sample Size",  n),
        ("Variables",    len(variables)),
        ("Significance", "* p<0.05  ** p<0.01  *** p<0.001"),
    ]:
        ws4.append(row)
    ws4.column_dimensions["A"].width = 20
    ws4.column_dimensions["B"].width = 40

    wb.save(out_path)


# ─── Public API ───────────────────────────────────────────────────────────────

def run_analysis(
    data: dict,
    method: str,
    params: dict,
    session_dir: Path,
) -> dict:
    """
    Full analysis: compute matrices, draw heatmap + scatter, write Excel.
    Returns result dict (same shape as the old R JSON output).
    """
    df = pd.DataFrame(data).select_dtypes(include="number").dropna(how="all")
    if df.shape[1] < 2:
        raise ValueError("Need at least 2 numeric columns")

    n         = int(df.shape[0])
    variables = list(df.columns)

    cor_mat, p_mat, sig_mat, variables = compute_matrices(df, method)

    # Draw plots
    _draw_heatmap(
        cor_mat, p_mat, variables, method, n, params,
        session_dir / "correlation_heatmap.png",
    )
    _draw_scatter(
        df, cor_mat, p_mat, variables, method, params,
        session_dir / "scatter_matrix.png",
    )

    # Write Excel
    _save_excel(
        cor_mat, p_mat, sig_mat, variables,
        method, n,
        session_dir / "correlation_results.xlsx",
    )

    return {
        "status":     "success",
        "method":     method,
        "n":          n,
        "variables":  variables,
        "cor_matrix": cor_mat.tolist(),
        "p_matrix":   p_mat.tolist(),
        "sig_matrix": [[str(v) for v in row] for row in sig_mat],
    }


def regenerate_heatmap(
    cor_mat_list: list,
    p_mat_list: list,
    variables: list,
    method: str,
    n: int,
    params: dict,
    session_dir: Path,
) -> None:
    """
    Re-draw heatmap only (called by the live-update endpoint).
    Mirrors heatmap_only.R behaviour exactly.
    """
    cor_mat = np.array(cor_mat_list)
    p_mat   = np.array(p_mat_list)

    _draw_heatmap(
        cor_mat, p_mat, variables, method, n, params,
        session_dir / "correlation_heatmap.png",
    )
