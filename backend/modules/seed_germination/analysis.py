"""
Seed Germination & Seedling Analysis Engine  v2
Germination: FGP, GSI, T10/50/90, UnifG, MGT, MGR, VarGer, CVt, CVG,
             Sinc, Unc, Timson, ERI, Weibull-T50
Seedling:    Shoot/Root/Total length, S:R ratio, Unif1, Unif2,
             Growth, Vigor_Unif1, Vigor_Unif2, Vigor_corr x2,
             STI (shoot/root/total DW), Shoot/Root/Total DW
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.stats import weibull_min
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

PALETTE = ["#0072B2","#D55E00","#009E73","#E69F00",
           "#CC79A7","#56B4E9","#737373","#1C1C1C"]
SHAPE_MAP = {"circle":"o","triangle":"^","square":"s","diamond":"D","cross":"P"}

# ══════════════════════════════════════════════════════════════
# GERMINATION INDICES
# ══════════════════════════════════════════════════════════════

def _daily(nger):
    n = [nger[0]]
    for i in range(1, len(nger)):
        n.append(max(0, nger[i] - nger[i-1]))
    return np.array(n, dtype=float)

def calc_FGP(nger, ns):    return float(max(nger)) / ns * 100
def calc_GSI(t, ng):
    n = _daily(ng)
    return float(sum(n[i]/t[i] for i in range(len(t)) if t[i]>0))
def calc_MGT(t, ng):
    n = _daily(ng); tot = sum(n)
    if tot == 0: return np.nan
    return float(sum(t[i]*n[i] for i in range(len(t))) / tot)
def calc_MGR(t, ng):
    m = calc_MGT(t, ng); return float(1/m) if m and m!=0 else np.nan

def _Tx(t, ng, pct):
    N = max(ng); tgt = N*pct/100
    t, ng = np.array(t), np.array(ng)
    ni, ti, nj, tj = N, max(t), 0, min(t)
    for i in range(len(t)-1,-1,-1):
        if ng[i] > tgt: ni, ti = ng[i], t[i]
    for j in range(len(t)):
        if ng[j] < tgt: nj, tj = ng[j], t[j]
    if ni == nj: return float(ti)
    return float(ti + (tgt-ni)*(tj-ti)/(nj-ni))

def calc_T10(t,ng):    return _Tx(t,ng,10)
def calc_T50(t,ng):    return _Tx(t,ng,50)
def calc_T90(t,ng):    return _Tx(t,ng,90)
def calc_UnifG(t,ng):  return calc_T90(t,ng)-calc_T10(t,ng)

def calc_VarGer(t,ng):
    n=_daily(ng); mgt=calc_MGT(t,ng); tot=sum(n)
    if tot<=1: return np.nan
    return float(sum(n[i]*(t[i]-mgt)**2 for i in range(len(n)))/(tot-1))
def calc_CVt(t,ng):
    mgt=calc_MGT(t,ng); var=calc_VarGer(t,ng)
    if not mgt or mgt==0 or var is None: return np.nan
    return float(np.sqrt(var)/mgt*100)
def calc_CVG(t,ng):
    n=_daily(ng)
    den=sum(n[i]*t[i] for i in range(len(n)))
    return float(sum(n)/den*100) if den!=0 else np.nan
def calc_Sinc(t,ng):
    n=_daily(ng); N=sum(n); N2=N*(N-1)/2
    if N2==0: return np.nan
    return float(sum(ni*(ni-1)/2 for ni in n)/N2)
def calc_Unc(t,ng):
    n=_daily(ng); tot=sum(n)
    if tot==0: return np.nan
    E=0
    for ni in n:
        f=ni/tot
        if f>0: E+=f*np.log2(f)
    return float(-E)
def calc_Timson(t,ng):
    N=max(ng)
    if N==0: return np.nan
    return float(sum(g/N*100 for g in ng))
def calc_ERI(t,ng):
    N=max(ng)
    if N==0: return np.nan
    n=_daily(ng)
    return float(sum(n[i]/N*100/t[i] for i in range(len(t)) if t[i]>0))

def calc_weibull_T50(t, ng):
    N=max(ng)
    if N==0: return np.nan, np.nan, np.nan, None
    prop = np.array(ng)/N
    t_arr = np.array(t, dtype=float)
    mask = prop>0
    if mask.sum()<2: return np.nan, np.nan, np.nan, None
    try:
        counts = np.round(prop[mask]*100).astype(int).clip(1)
        data = np.repeat(t_arr[mask], counts)
        sh, loc, sc = weibull_min.fit(data, floc=0)
        t50 = sc*(np.log(2))**(1/sh)
        tf = np.linspace(0, max(t_arr)*1.1, 300)
        fitted = weibull_min.cdf(tf, sh, loc=0, scale=sc)
        return float(t50), float(sh), float(sc), (tf, fitted)
    except Exception:
        return np.nan, np.nan, np.nan, None

def compute_germination_indices(gdf, nseeds):
    day_cols = [c for c in gdf.columns if c not in ("Treatment","Replicate")]
    days = np.array([float(c) for c in day_cols])
    results = []; weibull_fits = {}
    for _, row in gdf.iterrows():
        trt=row["Treatment"]; rep=row["Replicate"]
        ng=row[day_cols].values.astype(float)
        t50w,wsh,wsc,wcurve = calc_weibull_T50(days,ng)
        if wcurve is not None: weibull_fits[(trt,rep)]=wcurve
        results.append({
            "Treatment":trt,"Replicate":rep,
            "FGP":calc_FGP(ng,nseeds),"GSI":calc_GSI(days,ng),
            "T10":calc_T10(days,ng),"T50":calc_T50(days,ng),
            "T90":calc_T90(days,ng),"UnifG":calc_UnifG(days,ng),
            "MGT":calc_MGT(days,ng),"MGR":calc_MGR(days,ng),
            "VarGer":calc_VarGer(days,ng),"CVt":calc_CVt(days,ng),
            "CVG":calc_CVG(days,ng),"Sinc":calc_Sinc(days,ng),
            "Unc":calc_Unc(days,ng),"Timson":calc_Timson(days,ng),
            "ERI":calc_ERI(days,ng),"T50_Weibull":t50w,
            "Weibull_shape":wsh,"Weibull_scale":wsc,
        })
    return pd.DataFrame(results), weibull_fits

# ══════════════════════════════════════════════════════════════
# SEEDLING INDICES
# ══════════════════════════════════════════════════════════════

def calc_unif1(shoot, root):
    total = np.array(shoot)+np.array(root)
    alive = total[total>0]
    if len(alive)==0: return np.nan
    ndeads = len(total)-len(alive)
    desv = sum(abs(v-np.mean(total)) for v in alive)
    penalty = ndeads*(50/len(total))
    return float((1-desv/(len(alive)*np.mean(alive)))*1000-penalty)

def calc_unif2(shoot, root):
    sh,ro = np.array(shoot,float), np.array(root,float)
    tot = sh+ro
    razao = np.where(sh>0, ro/sh, 0.0)
    return float(1000-(0.75*np.std(sh)+0.5*np.std(ro)
                       +2.5*np.std(tot)+50*np.std(razao)))

def calc_growth(shoot, root, wr, wh):
    return float(np.nanmean(root)*wr + np.nanmean(shoot)*wh)

def compute_seedling_indices(sdf, control_trt, wr, wh, wg, wu):
    ctrl = sdf[sdf["Treatment"]==control_trt] if control_trt else pd.DataFrame()
    results=[]
    for (trt,rep), grp in sdf.groupby(["Treatment","Replicate"]):
        sh  = grp["Shoot_cm"].values.astype(float)
        ro  = grp["Root_cm"].values.astype(float)
        sdw = grp["Shoot_DW_g"].values.astype(float)
        rdw = grp["Root_DW_g"].values.astype(float)
        u1  = calc_unif1(sh, ro)
        u2  = calc_unif2(sh, ro)
        g   = calc_growth(sh, ro, wr, wh)
        v1  = g*wg + u1*wu
        v2  = g*wg + u2*wu
        def sti(v, ctrl_vals):
            mc = np.nanmean(ctrl_vals)
            return float((v*mc)/mc**2) if mc!=0 else np.nan
        row = {
            "Treatment":trt,"Replicate":rep,
            "Mean_Shoot_cm":float(np.nanmean(sh)),
            "Mean_Root_cm": float(np.nanmean(ro)),
            "Mean_Total_cm":float(np.nanmean(sh+ro)),
            "Shoot_Root_ratio":float(np.nanmean(ro)/np.nanmean(sh))
                               if np.nanmean(sh)!=0 else np.nan,
            "Unif1":u1,"Unif2":u2,
            "Growth":g,"Vigor_Unif1":v1,"Vigor_Unif2":v2,
            "Vigor_corr_Unif1":np.nan,"Vigor_corr_Unif2":np.nan,
            "Mean_Shoot_DW":float(np.nanmean(sdw)),
            "Mean_Root_DW": float(np.nanmean(rdw)),
            "Total_DW":     float(np.nanmean(sdw+rdw)),
            "DW_Ratio":     float(np.nanmean(rdw)/np.nanmean(sdw))
                            if np.nanmean(sdw)!=0 else np.nan,
            "STI_Shoot":    np.nan,"STI_Root":np.nan,"STI_Total_DW":np.nan,
        }
        if trt!=control_trt and len(ctrl)>0:
            row["STI_Shoot"]    = sti(np.nanmean(sh),    ctrl["Shoot_cm"].values)
            row["STI_Root"]     = sti(np.nanmean(ro),    ctrl["Root_cm"].values)
            row["STI_Total_DW"] = sti(np.nanmean(sdw+rdw),
                                      ctrl["Shoot_DW_g"].values+ctrl["Root_DW_g"].values)
        results.append(row)
    return pd.DataFrame(results)

# ══════════════════════════════════════════════════════════════
# PLOTS
# ══════════════════════════════════════════════════════════════

def _ax(ax, grid=True):
    ax.set_facecolor("#FAFAFA")
    ax.spines[["top","right"]].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#999999",labelsize=9)
    ax.grid(grid,linestyle="--",linewidth=0.45,color="#DDDDDD",alpha=0.8,zorder=0)

def _save(fig, path, name):
    fig.savefig(path/f"{name}.png",dpi=150,bbox_inches="tight",facecolor="#FAFAFA")
    fig.savefig(path/f"{name}.svg",format="svg",bbox_inches="tight",facecolor="#FAFAFA")
    plt.close(fig)

def plot_germination_curve(gdf, wfits, params, sdir):
    day_cols=[c for c in gdf.columns if c not in ("Treatment","Replicate")]
    days=np.array([float(c) for c in day_cols])
    trts=gdf["Treatment"].unique()
    grid=bool(params.get("show_grid",True))
    title=params.get("title_gcurve","Germination Curve")
    pal=PALETTE; mk=SHAPE_MAP.get(params.get("point_shape","circle"),"o")
    sz=float(params.get("point_size",6))

    fig,ax=plt.subplots(figsize=(8,5.5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    for idx,trt in enumerate(trts):
        col=pal[idx%len(pal)]
        sub=gdf[gdf["Treatment"]==trt]
        N=sub[day_cols].values.astype(float).max()
        if N==0: continue
        mean_pct=sub[day_cols].values.astype(float).mean(axis=0)/N*100
        ax.plot(days,mean_pct,color=col,linewidth=2,zorder=3,label=trt)
        ax.scatter(days,mean_pct,color=col,s=sz**2,marker=mk,
                   edgecolors="white",linewidths=0.5,zorder=4)
        curves=[wfits[k] for k in wfits if k[0]==trt]
        if curves:
            tf=curves[0][0]
            mf=np.mean([c[1] for c in curves],axis=0)*100
            ax.plot(tf,mf,color=col,linewidth=1.2,linestyle="--",alpha=0.7,zorder=2)
    ax.set_xlabel("Days after sowing",fontsize=10.5,labelpad=7,color="#333")
    ax.set_ylabel("Cumulative Germination (%)",fontsize=10.5,labelpad=7,color="#333")
    ax.set_ylim(0,105)
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    from matplotlib.lines import Line2D
    leg1=ax.legend(loc="upper left",frameon=True,framealpha=0.88,
                   fontsize=8.5,edgecolor="#DDDDDD")
    ax.add_artist(leg1)
    ax.legend(handles=[Line2D([0],[0],color="grey",lw=2,label="Observed (mean)"),
                       Line2D([0],[0],color="grey",lw=1.2,ls="--",label="Weibull fit")],
              loc="lower right",fontsize=8,frameon=True,framealpha=0.88,edgecolor="#DDDDDD")
    fig.tight_layout(pad=1.4); _save(fig,sdir,"germination_curve")

def plot_daily_rate(gdf, params, sdir):
    day_cols=[c for c in gdf.columns if c not in ("Treatment","Replicate")]
    days=[float(c) for c in day_cols]
    trts=gdf["Treatment"].unique(); n_trt=len(trts)
    grid=bool(params.get("show_grid",True))
    title=params.get("title_rate","Daily Germination Rate")
    pal=PALETTE; width=0.8/n_trt; x=np.arange(len(days))
    fig,ax=plt.subplots(figsize=(9,5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    for idx,trt in enumerate(trts):
        sub=gdf[gdf["Treatment"]==trt]
        cum=sub[day_cols].values.astype(float).mean(axis=0)
        daily=np.diff(np.concatenate([[0],cum]))
        ax.bar(x+idx*width-(n_trt-1)*width/2,daily,
               width=width*0.85,color=pal[idx%len(pal)],alpha=0.82,label=trt,zorder=3)
    ax.set_xticks(x); ax.set_xticklabels([f"Day {int(d)}" for d in days],
                                          fontsize=8.5,rotation=30)
    ax.set_xlabel("Time",fontsize=10.5,labelpad=7,color="#333")
    ax.set_ylabel("Mean daily germination (seeds)",fontsize=10.5,labelpad=7,color="#333")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    ax.legend(frameon=True,framealpha=0.88,fontsize=8.5,edgecolor="#DDDDDD")
    fig.tight_layout(pad=1.4); _save(fig,sdir,"daily_rate")

def plot_index_comparison(gres, params, sdir):
    indices=["FGP","GSI","MGT","MGR","T50","T50_Weibull",
             "UnifG","CVt","CVG","Timson","ERI","Sinc"]
    grid=bool(params.get("show_grid",True))
    title=params.get("title_index","Germination Index Comparison")
    pal=PALETTE
    means=gres.groupby("Treatment")[indices].mean().reset_index()
    n_trt=len(means)
    fig,axes=plt.subplots(3,4,figsize=(14,9),dpi=130)
    fig.patch.set_facecolor("#FAFAFA"); axes=axes.flatten()
    for i,idx in enumerate(indices):
        ax=axes[i]; _ax(ax,grid)
        vals=means[idx].values
        cols=[pal[j%len(pal)] for j in range(n_trt)]
        bars=ax.bar(means["Treatment"],vals,color=cols,alpha=0.84,zorder=3)
        ax.set_title(idx,fontsize=9,fontweight=700,color="#1C1C1C",pad=5)
        ax.tick_params(axis="x",labelsize=7.5,rotation=20)
        for bar,v in zip(bars,vals):
            if not np.isnan(v):
                ax.text(bar.get_x()+bar.get_width()/2,
                        bar.get_height()+0.01*max(list(vals)+[1]),
                        f"{v:.2f}",ha="center",va="bottom",fontsize=7,color="#555")
    for j in range(len(indices),len(axes)): axes[j].set_visible(False)
    if title: fig.suptitle(title,fontsize=12,fontweight="bold",color="#1C1C1C",y=1.01)
    fig.tight_layout(pad=1.2); _save(fig,sdir,"index_comparison")

def plot_seedling_boxplots(sdf, params, sdir):
    grid=bool(params.get("show_grid",True))
    title=params.get("title_box","Seedling Length Distribution")
    pal=PALETTE; trts=sdf["Treatment"].unique()
    sdf=sdf.copy(); sdf["Total_cm"]=sdf["Shoot_cm"]+sdf["Root_cm"]
    fig,axes=plt.subplots(1,3,figsize=(13,5.5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA")
    for ax,(col,ylabel) in zip(axes,[("Shoot_cm","Shoot Length (cm)"),
                                      ("Root_cm","Root Length (cm)"),
                                      ("Total_cm","Total Length (cm)")]):
        _ax(ax,grid)
        data=[sdf[sdf["Treatment"]==t][col].dropna().values for t in trts]
        bp=ax.boxplot(data,patch_artist=True,notch=False,
                      medianprops=dict(color="#1C1C1C",linewidth=2))
        for patch,c in zip(bp["boxes"],pal): patch.set_facecolor(c); patch.set_alpha(0.75)
        ax.set_xticklabels(trts,fontsize=8.5,rotation=20)
        ax.set_ylabel(ylabel,fontsize=9.5,labelpad=6,color="#333")
        ax.set_xlabel("Treatment",fontsize=9,labelpad=5,color="#333")
    if title: fig.suptitle(title,fontsize=11.5,fontweight="bold",color="#1C1C1C")
    fig.tight_layout(pad=1.4); _save(fig,sdir,"seedling_boxplots")

def plot_dryweight_bar(sdf, params, sdir):
    grid=bool(params.get("show_grid",True))
    title=params.get("title_dw","Seedling Dry Weight by Treatment")
    pal=PALETTE; trts=sdf["Treatment"].unique()
    m=sdf.groupby("Treatment")[["Shoot_DW_g","Root_DW_g"]].mean().reset_index()
    m["Total_DW_g"]=m["Shoot_DW_g"]+m["Root_DW_g"]
    x=np.arange(len(trts)); width=0.25
    fig,ax=plt.subplots(figsize=(8.5,5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    for i,(col,lbl) in enumerate([("Shoot_DW_g","Shoot DW"),
                                    ("Root_DW_g","Root DW"),
                                    ("Total_DW_g","Total DW")]):
        vals=m.set_index("Treatment").reindex(trts)[col].values
        bars=ax.bar(x+i*width-width,vals,width=width*0.85,
                    color=pal[i],alpha=0.84,label=lbl,zorder=3)
        for bar,v in zip(bars,vals):
            if not np.isnan(v):
                ax.text(bar.get_x()+bar.get_width()/2,bar.get_height()+0.0005,
                        f"{v:.4f}",ha="center",va="bottom",fontsize=7,color="#555")
    ax.set_xticks(x); ax.set_xticklabels(trts,fontsize=9.5,rotation=15)
    ax.set_ylabel("Dry weight (g)",fontsize=10.5,labelpad=7,color="#333")
    ax.set_xlabel("Treatment",fontsize=10.5,labelpad=7,color="#333")
    ax.legend(frameon=True,framealpha=0.88,fontsize=8.5,edgecolor="#DDDDDD")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    fig.tight_layout(pad=1.4); _save(fig,sdir,"dryweight_bar")

# ══════════════════════════════════════════════════════════════
# EXCEL
# ══════════════════════════════════════════════════════════════

def save_excel(gres, sres, sdir):
    wb=openpyxl.Workbook()
    hf=PatternFill("solid",fgColor="1C3557")
    hfont=Font(bold=True,color="FFFFFF",size=10)
    ctr=Alignment(horizontal="center")
    def hr(ws,vals):
        ws.append(vals)
        for c in ws[ws.max_row]: c.font=hfont; c.fill=hf; c.alignment=ctr
    ws1=wb.active; ws1.title="Germination Indices"
    if gres is not None:
        hr(ws1,list(gres.columns))
        for _,row in gres.iterrows():
            ws1.append([round(v,4) if isinstance(v,float) else v for v in row.values])
        for i,c in enumerate(gres.columns,1):
            ws1.column_dimensions[openpyxl.utils.get_column_letter(i)].width=max(14,len(c)+2)
    else: ws1.append(["No germination data provided"])
    ws2=wb.create_sheet("Seedling Metrics")
    if sres is not None:
        hr(ws2,list(sres.columns))
        for _,row in sres.iterrows():
            ws2.append([round(v,4) if isinstance(v,float) else v for v in row.values])
        for i,c in enumerate(sres.columns,1):
            ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width=max(16,len(c)+2)
    else: ws2.append(["No seedling data provided"])
    wb.save(sdir/"seed_results.xlsx")

# ══════════════════════════════════════════════════════════════
# PUBLIC API
# ══════════════════════════════════════════════════════════════

def run_analysis(germ_data, seed_data, nseeds, control_trt, params, sdir):
    gres=None; sres=None; wfits={}; plots=[]
    if germ_data is not None:
        gdf=pd.DataFrame(germ_data)
        gres,wfits=compute_germination_indices(gdf,nseeds)
        plot_germination_curve(gdf,wfits,params,sdir); plots.append("gcurve")
        plot_daily_rate(gdf,params,sdir); plots.append("rate")
        plot_index_comparison(gres,params,sdir); plots.append("index")
    if seed_data is not None:
        sdf=pd.DataFrame(seed_data)
        sres=compute_seedling_indices(
            sdf, control_trt,
            wr=float(params.get("wr",90)), wh=float(params.get("wh",10)),
            wg=float(params.get("wg",0.7)), wu=float(params.get("wu",0.3)))
        if gres is not None:
            fgp_map=gres.groupby("Treatment")["FGP"].mean().to_dict()
            sres["Vigor_corr_Unif1"]=sres.apply(
                lambda r: r["Vigor_Unif1"]*fgp_map.get(r["Treatment"],100)/100, axis=1)
            sres["Vigor_corr_Unif2"]=sres.apply(
                lambda r: r["Vigor_Unif2"]*fgp_map.get(r["Treatment"],100)/100, axis=1)
        plot_seedling_boxplots(sdf,params,sdir); plots.append("box")
        plot_dryweight_bar(sdf,params,sdir); plots.append("dw")
    save_excel(gres,sres,sdir)
    trts=[]
    if germ_data: trts=pd.DataFrame(germ_data)["Treatment"].unique().tolist()
    elif seed_data: trts=pd.DataFrame(seed_data)["Treatment"].unique().tolist()
    return {
        "germ_results": gres.to_dict(orient="records") if gres is not None else None,
        "seed_results": sres.to_dict(orient="records") if sres is not None else None,
        "plots": plots, "treatments": trts,
    }

def regenerate_plots(stored, params, sdir):
    if stored.get("germ_data"):
        gdf=pd.DataFrame(stored["germ_data"])
        _,wf=compute_germination_indices(gdf,stored.get("nseeds",100))
        plot_germination_curve(gdf,wf,params,sdir)
        plot_daily_rate(gdf,params,sdir)
        plot_index_comparison(pd.DataFrame(stored["germ_results"]),params,sdir)
    if stored.get("seed_data"):
        sdf=pd.DataFrame(stored["seed_data"])
        plot_seedling_boxplots(sdf,params,sdir)
        plot_dryweight_bar(sdf,params,sdir)
