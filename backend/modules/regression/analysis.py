"""
Regression Module — Analysis Engine
Supports: simple linear, multiple linear, stepwise (forward/backward/both)
Plots:    fitted-vs-observed, residuals-vs-fitted, Q-Q, coefficient forest
Export:   PNG + SVG plots, Excel (3 sheets)
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy import stats
import statsmodels.api as sm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

PALETTE = [
    "#0072B2","#D55E00","#009E73","#E69F00",
    "#CC79A7","#56B4E9","#737373","#1C1C1C",
]
SHAPE_MAP = {
    "circle":"o","triangle":"^","square":"s","diamond":"D","cross":"P",
}

# ─── Stepwise ─────────────────────────────────────────────────────────────────

def _fit(X_df, cols, y):
    Xc = sm.add_constant(X_df[cols].values, has_constant="add")
    return sm.OLS(y, Xc).fit()

def _score(m, criterion):
    return m.aic if criterion in ("aic","p_value") else m.bic

def _stepwise(X_df, y, direction, criterion, p_threshold):
    candidates = list(X_df.columns)
    selected   = []

    if direction in ("forward","both"):
        improved = True
        while improved and candidates:
            improved = False
            cur_score = _score(_fit(X_df, selected, y), criterion) if selected else 1e18
            best_col = None
            for col in candidates:
                trial = selected + [col]
                m = _fit(X_df, trial, y)
                if criterion == "p_value":
                    p = m.pvalues.get(col, m.pvalues.iloc[-1])
                    if p < p_threshold and _score(m, criterion) < cur_score:
                        cur_score = _score(m, criterion); best_col = col
                else:
                    s = _score(m, criterion)
                    if s < cur_score:
                        cur_score = s; best_col = col
            if best_col:
                selected.append(best_col); candidates.remove(best_col); improved = True

    elif direction == "backward":
        selected = list(candidates)
        improved = True
        while improved and len(selected) > 1:
            improved  = False
            cur_score = _score(_fit(X_df, selected, y), criterion)
            worst_col = None
            for col in selected:
                rem = [c for c in selected if c != col]
                if not rem: continue
                m = _fit(X_df, rem, y)
                if criterion == "p_value":
                    full_m = _fit(X_df, selected, y)
                    p = full_m.pvalues.get(col, full_m.pvalues.iloc[-1])
                    if p > p_threshold and _score(m, criterion) <= cur_score:
                        cur_score = _score(m, criterion); worst_col = col
                else:
                    s = _score(m, criterion)
                    if s < cur_score:
                        cur_score = s; worst_col = col
            if worst_col:
                selected.remove(worst_col); improved = True

    return selected if selected else [X_df.columns[0]]

# ─── Public API ───────────────────────────────────────────────────────────────

def run_regression(data, y_col, x_cols, reg_type, stepwise_dir,
                   stepwise_criterion, stepwise_threshold, params, session_dir):
    df  = pd.DataFrame(data)
    sub = df[[y_col] + x_cols].dropna()
    y   = sub[y_col].values.astype(float)
    X_df= sub[x_cols].astype(float)
    n   = len(y)
    if n < 4:
        raise ValueError(f"Need at least 4 complete rows (got {n})")

    if reg_type == "simple":
        final_cols = x_cols[:1]
    elif reg_type == "stepwise":
        final_cols = _stepwise(X_df, y, stepwise_dir,
                               stepwise_criterion, stepwise_threshold)
    else:
        final_cols = x_cols

    X_mat = sm.add_constant(X_df[final_cols].values, has_constant="add")
    model = sm.OLS(y, X_mat).fit()

    labels = ["Intercept"] + final_cols
    ci     = model.conf_int(alpha=0.05)
    coef_table = [
        {"variable": lbl,
         "beta":     float(model.params[i]),
         "se":       float(model.bse[i]),
         "t":        float(model.tvalues[i]),
         "p":        float(model.pvalues[i]),
         "ci_lower": float(ci[0][i]),
         "ci_upper": float(ci[1][i])}
        for i, lbl in enumerate(labels)
    ]

    fitted    = model.fittedvalues
    residuals = model.resid
    fit_stats = {
        "r2":     float(model.rsquared),
        "adj_r2": float(model.rsquared_adj),
        "rmse":   float(np.sqrt(np.mean(residuals**2))),
        "f_stat": float(model.fvalue)   if model.fvalue   is not None else None,
        "f_pval": float(model.f_pvalue) if model.f_pvalue is not None else None,
        "n": int(n), "k": len(final_cols),
        "aic": float(model.aic), "bic": float(model.bic),
    }

    ss_res = float(np.sum(residuals**2))
    ss_tot = float(np.sum((y - np.mean(y))**2))
    ss_reg = ss_tot - ss_res
    df_reg = len(final_cols); df_res = n - df_reg - 1
    ms_reg = ss_reg/df_reg if df_reg > 0 else 0
    ms_res = ss_res/df_res if df_res > 0 else 0
    f_val  = ms_reg/ms_res if ms_res > 0 else 0
    f_p    = float(stats.f.sf(f_val, df_reg, df_res))
    anova  = [
        {"source":"Regression","ss":ss_reg,"df":df_reg,"ms":ms_reg,"f":f_val,"p":f_p},
        {"source":"Residual",  "ss":ss_res,"df":df_res,"ms":ms_res,"f":None, "p":None},
        {"source":"Total",     "ss":ss_tot,"df":n-1,   "ms":None,  "f":None, "p":None},
    ]

    result = {
        "y_col":y_col,"final_cols":final_cols,"reg_type":reg_type,
        "coef_table":coef_table,"fit_stats":fit_stats,"anova":anova,
        "y_actual":y.tolist(),"y_fitted":fitted.tolist(),
        "residuals":residuals.tolist(),
    }
    _draw_all(result, params, session_dir)
    _save_excel(result, session_dir)
    return result

def regenerate_plots(stored, params, session_dir):
    _draw_all(stored, params, session_dir)

# ─── Shared helpers ───────────────────────────────────────────────────────────

def _ax(ax, grid=True):
    ax.set_facecolor("#FAFAFA")
    ax.spines[["top","right"]].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#999999",labelsize=9)
    ax.grid(grid,linestyle="--",linewidth=0.45,color="#DDDDDD",alpha=0.8,zorder=0)

def _save(fig, session_dir, name):
    fig.savefig(session_dir/f"{name}.png",dpi=150,bbox_inches="tight",
                facecolor=fig.get_facecolor())
    fig.savefig(session_dir/f"{name}.svg",format="svg",bbox_inches="tight",
                facecolor=fig.get_facecolor())
    plt.close(fig)

def _draw_all(r, params, session_dir):
    _plot_fvo(r,params,session_dir)
    _plot_rvf(r,params,session_dir)
    _plot_qq(r,params,session_dir)
    _plot_forest(r,params,session_dir)

# ─── 1. Fitted vs Observed ────────────────────────────────────────────────────

def _plot_fvo(r,params,session_dir):
    y_act = np.array(r["y_actual"]); y_fit = np.array(r["y_fitted"])
    col   = params.get("point_color",PALETTE[0])
    mk    = SHAPE_MAP.get(params.get("point_shape","circle"),"o")
    sz    = float(params.get("point_size",6))
    grid  = bool(params.get("show_grid",True))
    title = params.get("title_fvo","Fitted vs Observed")

    fig,ax = plt.subplots(figsize=(6.5,5.5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    ax.scatter(y_act,y_fit,c=col,s=sz**2,marker=mk,alpha=0.8,
               edgecolors="white",linewidths=0.5,zorder=3)
    lo=min(y_act.min(),y_fit.min()); hi=max(y_act.max(),y_fit.max())
    pad=(hi-lo)*0.05
    ax.plot([lo-pad,hi+pad],[lo-pad,hi+pad],color="#CCCCCC",
            linewidth=1.5,linestyle="--",zorder=2)
    ax.text(0.05,0.95,f"R\u00b2 = {r['fit_stats']['r2']:.4f}",
            transform=ax.transAxes,fontsize=9,va="top",color="#555555",
            bbox=dict(boxstyle="round,pad=0.35",fc="white",ec="#DDDDDD",alpha=0.88))
    ax.set_xlabel(f"Observed ({r['y_col']})",fontsize=10.5,labelpad=7,color="#333333")
    ax.set_ylabel(f"Fitted ({r['y_col']})",  fontsize=10.5,labelpad=7,color="#333333")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    fig.tight_layout(pad=1.4); _save(fig,session_dir,"fitted_vs_observed")

# ─── 2. Residuals vs Fitted ───────────────────────────────────────────────────

def _plot_rvf(r,params,session_dir):
    y_fit = np.array(r["y_fitted"]); resid = np.array(r["residuals"])
    col   = params.get("point_color",PALETTE[0])
    mk    = SHAPE_MAP.get(params.get("point_shape","circle"),"o")
    sz    = float(params.get("point_size",6))
    grid  = bool(params.get("show_grid",True))
    title = params.get("title_rvf","Residuals vs Fitted")

    fig,ax = plt.subplots(figsize=(6.5,5.0),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    ax.scatter(y_fit,resid,c=col,s=sz**2,marker=mk,alpha=0.8,
               edgecolors="white",linewidths=0.5,zorder=3)
    ax.axhline(0,color="#AAAAAA",linewidth=1.2,zorder=2)
    try:
        from statsmodels.nonparametric.smoothers_lowess import lowess
        lo=lowess(resid,y_fit,frac=0.6,return_sorted=True)
        ax.plot(lo[:,0],lo[:,1],color=PALETTE[1],linewidth=1.6,zorder=4)
    except Exception:
        pass
    ax.set_xlabel("Fitted values",fontsize=10.5,labelpad=7,color="#333333")
    ax.set_ylabel("Residuals",    fontsize=10.5,labelpad=7,color="#333333")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    fig.tight_layout(pad=1.4); _save(fig,session_dir,"residuals_vs_fitted")

# ─── 3. Q-Q ──────────────────────────────────────────────────────────────────

def _plot_qq(r,params,session_dir):
    resid = np.array(r["residuals"])
    col   = params.get("point_color",PALETTE[0])
    mk    = SHAPE_MAP.get(params.get("point_shape","circle"),"o")
    sz    = float(params.get("point_size",6))
    grid  = bool(params.get("show_grid",True))
    title = params.get("title_qq","Normal Q-Q Plot")

    fig,ax = plt.subplots(figsize=(6.0,5.5),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)
    (osm,osr),(slope,intercept,_) = stats.probplot(resid,dist="norm")
    ax.scatter(osm,osr,c=col,s=sz**2,marker=mk,alpha=0.8,
               edgecolors="white",linewidths=0.5,zorder=3)
    xl=np.array([osm[0],osm[-1]])
    ax.plot(xl,slope*xl+intercept,color="#CCCCCC",linewidth=1.5,
            linestyle="--",zorder=2)
    ax.set_xlabel("Theoretical quantiles",fontsize=10.5,labelpad=7,color="#333333")
    ax.set_ylabel("Sample quantiles",     fontsize=10.5,labelpad=7,color="#333333")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    fig.tight_layout(pad=1.4); _save(fig,session_dir,"qq_plot")

# ─── 4. Coefficient forest ────────────────────────────────────────────────────

def _plot_forest(r,params,session_dir):
    coefs = [c for c in r["coef_table"] if c["variable"]!="Intercept"] or r["coef_table"]
    grid  = bool(params.get("show_grid",True))
    title = params.get("title_coef","Coefficient Plot (95% CI)")

    labels=[c["variable"] for c in coefs]; betas=[c["beta"] for c in coefs]
    ci_lo =[c["ci_lower"] for c in coefs]; ci_hi=[c["ci_upper"] for c in coefs]
    pvals =[c["p"] for c in coefs]

    n=len(coefs); fh=max(3.5,n*0.55+1.8)
    fig,ax=plt.subplots(figsize=(7.0,fh),dpi=150)
    fig.patch.set_facecolor("#FAFAFA"); _ax(ax,grid)

    for i,(b,lo,hi,p) in enumerate(zip(betas,ci_lo,ci_hi,pvals)):
        col=PALETTE[0] if p<0.05 else PALETTE[6]
        ax.plot([lo,hi],[i,i],color=col,linewidth=2.2,solid_capstyle="round",zorder=3)
        ax.scatter([b],[i],color=col,s=60,zorder=4,marker="D")

    ax.axvline(0,color="#CCCCCC",linewidth=1.2,linestyle="--",zorder=1)
    ax.set_yticks(np.arange(n)); ax.set_yticklabels(labels,fontsize=9.5)
    ax.set_xlabel("Coefficient (\u03b2)",fontsize=10.5,labelpad=7,color="#333333")

    from matplotlib.lines import Line2D
    ax.legend(handles=[
        Line2D([0],[0],marker="D",color="w",markerfacecolor=PALETTE[0],markersize=7,label="p < 0.05"),
        Line2D([0],[0],marker="D",color="w",markerfacecolor=PALETTE[6],markersize=7,label="p \u2265 0.05"),
    ],fontsize=8.5,frameon=True,framealpha=0.88,edgecolor="#DDDDDD")
    if title: ax.set_title(title,fontsize=11.5,fontweight="bold",color="#1C1C1C",pad=10)
    fig.tight_layout(pad=1.4); _save(fig,session_dir,"coef_forest")

# ─── Excel ────────────────────────────────────────────────────────────────────

def _save_excel(r,session_dir):
    wb=openpyxl.Workbook()
    hf=PatternFill("solid",fgColor="1C3557"); hfont=Font(bold=True,color="FFFFFF",size=10)
    sf=PatternFill("solid",fgColor="EAF3FB"); sfont=Font(bold=True,color="1C3557",size=9)
    ctr=Alignment(horizontal="center")

    def hr(ws,vals):
        ws.append(vals)
        for c in ws[ws.max_row]: c.font=hfont; c.fill=hf; c.alignment=ctr

    def stars(p):
        if p is None: return ""
        return "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns"

    ws1=wb.active; ws1.title="Coefficients"
    hr(ws1,["Variable","β (Coef)","Std Error","t-value","p-value","Sig","CI Lower","CI Upper"])
    for c in r["coef_table"]:
        ws1.append([c["variable"],round(c["beta"],6),round(c["se"],6),
                    round(c["t"],4),round(c["p"],6),stars(c["p"]),
                    round(c["ci_lower"],6),round(c["ci_upper"],6)])
    for col,w in zip("ABCDEFGH",[22,14,14,12,12,6,14,14]):
        ws1.column_dimensions[col].width=w

    ws2=wb.create_sheet("Model Fit")
    hr(ws2,["Metric","Value"])
    fs=r["fit_stats"]
    for lbl,val in [("R²",round(fs["r2"],6)),("Adjusted R²",round(fs["adj_r2"],6)),
                    ("RMSE",round(fs["rmse"],6)),
                    ("F-statistic",round(fs["f_stat"],4) if fs["f_stat"] else "—"),
                    ("F p-value",round(fs["f_pval"],6) if fs["f_pval"] else "—"),
                    ("AIC",round(fs["aic"],3)),("BIC",round(fs["bic"],3)),
                    ("n",fs["n"]),("k (predictors)",fs["k"]),
                    ("Type",r["reg_type"]),
                    ("Y",r["y_col"]),("Predictors",", ".join(r["final_cols"]))]:
        ws2.append([lbl,val])
    ws2.column_dimensions["A"].width=24; ws2.column_dimensions["B"].width=32

    ws3=wb.create_sheet("ANOVA")
    hr(ws3,["Source","SS","df","MS","F","p-value","Sig"])
    for row in r["anova"]:
        ws3.append([row["source"],round(row["ss"],6),row["df"],
                    round(row["ms"],6) if row["ms"] else "—",
                    round(row["f"],4)  if row["f"]  else "—",
                    round(row["p"],6)  if row["p"]  else "—",
                    stars(row["p"])    if row["p"]  else ""])
    for col,w in zip("ABCDEFG",[14,16,8,16,12,12,6]):
        ws3.column_dimensions[col].width=w

    wb.save(session_dir/"regression_results.xlsx")
