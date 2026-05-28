"""
PCA Module — Analysis Engine
Handles PCA computation and all four plot types:
  scree · score · biplot · loading heatmap
"""

import json
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from matplotlib.patches import FancyArrowPatch
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

# ─── Palette (colorblind-safe, mirrors TwoVarCorrelation swatches) ────────────

PALETTE = [
    "#0072B2",  # Steel Blue
    "#D55E00",  # Vermillion
    "#009E73",  # Forest Green
    "#E69F00",  # Orange
    "#CC79A7",  # Pink
    "#56B4E9",  # Sky Blue
    "#F0E442",  # Yellow
    "#1C1C1C",  # Black
]

SHAPE_MAP = {
    "circle":   "o",
    "triangle": "^",
    "square":   "s",
    "diamond":  "D",
    "cross":    "P",
}

# ─── Public API ───────────────────────────────────────────────────────────────

def run_pca(
    data: dict,
    selected_cols: list,
    group_col: str | None,
    group_data: list | None,
    params: dict,
    session_dir: Path,
) -> dict:
    """
    Fit PCA, generate all four plots, write Excel, return result dict.
    The result dict contains everything needed for regenerate_plots.
    """
    X_df = {c: data[c] for c in selected_cols}
    X    = np.array([X_df[c] for c in selected_cols], dtype=float).T  # shape (n, p)

    # Drop rows where any selected col is NaN
    mask = ~np.isnan(X).any(axis=1)
    X    = X[mask]
    n, p = X.shape

    if n < 3:
        raise ValueError(f"Need ≥ 3 complete rows (got {n} after dropping NaN rows)")
    if p < 2:
        raise ValueError("Need ≥ 2 selected columns")

    # Group labels (after same NaN mask)
    groups = None
    if group_col and group_data:
        g_arr  = np.array(group_data)
        groups = list(g_arr[mask])

    # Scale
    scale = bool(params.get("scale", True))
    if scale:
        scaler = StandardScaler()
        X_fit  = scaler.fit_transform(X)
    else:
        X_fit = X.copy()

    # Fit PCA
    n_components = min(int(params.get("n_components", 5)), n, p)
    pca          = PCA(n_components=n_components)
    scores       = pca.fit_transform(X_fit)        # (n, k)
    loadings     = pca.components_.T               # (p, k)  variables × PCs
    explained    = pca.explained_variance_ratio_   # (k,)
    cumulative   = np.cumsum(explained)

    pc_labels = [f"PC{i+1}" for i in range(n_components)]

    result = {
        "n":           int(n),
        "p":           int(p),
        "n_components": int(n_components),
        "selected_cols": selected_cols,
        "group_col":   group_col,
        "groups":      groups,
        "explained":   explained.tolist(),
        "cumulative":  cumulative.tolist(),
        "pc_labels":   pc_labels,
        "scores":      scores.tolist(),
        "loadings":    loadings.tolist(),
        # Raw data stored so replot never needs the uploaded file
        "X_fit":       X_fit.tolist(),
    }

    _draw_all(result, params, session_dir)
    _save_excel(result, session_dir)

    return result


def regenerate_plots(stored: dict, params: dict, session_dir: Path) -> None:
    _draw_all(stored, params, session_dir)


# ─── Plot dispatcher ──────────────────────────────────────────────────────────

def _draw_all(r: dict, params: dict, session_dir: Path) -> None:
    _plot_scree(r, params, session_dir)
    _plot_scores(r, params, session_dir)
    _plot_biplot(r, params, session_dir)
    _plot_loadings_heatmap(r, params, session_dir)


# ─── Shared style helpers ─────────────────────────────────────────────────────

def _base_ax(ax, show_grid=True):
    ax.set_facecolor("#FAFAFA")
    ax.spines[["top", "right"]].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#999999", labelsize=9)
    ax.grid(show_grid, linestyle="--", linewidth=0.45,
            color="#DDDDDD", alpha=0.8, zorder=0)


def _save_fig(fig, session_dir, name):
    fig.savefig(session_dir / f"{name}.png", dpi=150,
                bbox_inches="tight", facecolor=fig.get_facecolor())
    fig.savefig(session_dir / f"{name}.svg", format="svg",
                bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


# ─── 1. Scree plot ────────────────────────────────────────────────────────────

def _plot_scree(r: dict, params: dict, session_dir: Path) -> None:
    explained  = r["explained"]
    cumulative = r["cumulative"]
    pc_labels  = r["pc_labels"]
    show_grid  = bool(params.get("show_grid", True))
    title      = params.get("scree_title", "Scree Plot")

    fig, ax = plt.subplots(figsize=(7, 4.8), dpi=150)
    fig.patch.set_facecolor("#FAFAFA")
    _base_ax(ax, show_grid)

    x = np.arange(len(explained))
    ax.bar(x, [v * 100 for v in explained],
           color=PALETTE[0], alpha=0.82, zorder=3, width=0.55, label="Variance %")
    ax2 = ax.twinx()
    ax2.plot(x, [v * 100 for v in cumulative],
             color=PALETTE[1], linewidth=2.2, marker="o",
             markersize=6, zorder=4, label="Cumulative %")
    ax2.axhline(80, color="#DDDDDD", linewidth=1, linestyle=":")
    ax2.set_ylabel("Cumulative Variance (%)", fontsize=10, color=PALETTE[1], labelpad=7)
    ax2.tick_params(colors="#999999", labelsize=9)
    ax2.spines[["top", "right"]].set_color("#CCCCCC")
    ax2.set_ylim(0, 105)

    ax.set_xticks(x)
    ax.set_xticklabels(pc_labels, fontsize=9)
    ax.set_xlabel("Principal Component", fontsize=10.5, labelpad=7, color="#333333")
    ax.set_ylabel("Variance Explained (%)", fontsize=10.5, labelpad=7, color="#333333")
    if title:
        ax.set_title(title, fontsize=11.5, fontweight="bold", color="#1C1C1C", pad=10)

    # Annotate bars
    for xi, val in zip(x, explained):
        ax.text(xi, val * 100 + 0.8, f"{val*100:.1f}%",
                ha="center", va="bottom", fontsize=8, color="#555555")

    fig.tight_layout(pad=1.4)
    _save_fig(fig, session_dir, "scree_plot")


# ─── 2. Score plot ────────────────────────────────────────────────────────────

def _plot_scores(r: dict, params: dict, session_dir: Path) -> None:
    scores     = np.array(r["scores"])
    groups     = r.get("groups")
    explained  = r["explained"]
    pc_x_idx   = int(params.get("pc_x", 1)) - 1
    pc_y_idx   = int(params.get("pc_y", 2)) - 1
    pc_x_idx   = min(pc_x_idx, r["n_components"] - 1)
    pc_y_idx   = min(pc_y_idx, r["n_components"] - 1)
    show_grid  = bool(params.get("show_grid", True))
    title      = params.get("score_title", "PCA Score Plot")
    point_size = float(params.get("point_size", 6))
    point_shape= params.get("point_shape", "circle")
    marker     = SHAPE_MAP.get(point_shape, "o")

    pc_x_lab = f"PC{pc_x_idx+1} ({explained[pc_x_idx]*100:.1f}%)"
    pc_y_lab = f"PC{pc_y_idx+1} ({explained[pc_y_idx]*100:.1f}%)"

    fig, ax = plt.subplots(figsize=(7.2, 5.6), dpi=150)
    fig.patch.set_facecolor("#FAFAFA")
    _base_ax(ax, show_grid)

    if groups:
        unique_groups = list(dict.fromkeys(groups))   # preserve order, deduplicate
        color_map     = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(unique_groups)}
        for g in unique_groups:
            idx = [i for i, grp in enumerate(groups) if grp == g]
            ax.scatter(scores[idx, pc_x_idx], scores[idx, pc_y_idx],
                       c=color_map[g], s=point_size**2, marker=marker,
                       alpha=0.82, edgecolors="white", linewidths=0.55,
                       label=str(g), zorder=3)
        ax.legend(frameon=True, framealpha=0.88, fontsize=8.5,
                  edgecolor="#DDDDDD", loc="best")
    else:
        ax.scatter(scores[:, pc_x_idx], scores[:, pc_y_idx],
                   c=PALETTE[0], s=point_size**2, marker=marker,
                   alpha=0.82, edgecolors="white", linewidths=0.55, zorder=3)

    ax.axhline(0, color="#CCCCCC", linewidth=0.8, zorder=1)
    ax.axvline(0, color="#CCCCCC", linewidth=0.8, zorder=1)
    ax.set_xlabel(pc_x_lab, fontsize=10.5, labelpad=7, color="#333333")
    ax.set_ylabel(pc_y_lab, fontsize=10.5, labelpad=7, color="#333333")
    if title:
        ax.set_title(title, fontsize=11.5, fontweight="bold", color="#1C1C1C", pad=10)

    fig.tight_layout(pad=1.4)
    _save_fig(fig, session_dir, "score_plot")


# ─── 3. Biplot ────────────────────────────────────────────────────────────────

def _plot_biplot(r: dict, params: dict, session_dir: Path) -> None:
    scores      = np.array(r["scores"])
    loadings    = np.array(r["loadings"])
    explained   = r["explained"]
    sel_cols    = r["selected_cols"]
    groups      = r.get("groups")
    pc_x_idx    = int(params.get("pc_x", 1)) - 1
    pc_y_idx    = int(params.get("pc_y", 2)) - 1
    pc_x_idx    = min(pc_x_idx, r["n_components"] - 1)
    pc_y_idx    = min(pc_y_idx, r["n_components"] - 1)
    show_grid   = bool(params.get("show_grid", True))
    title       = params.get("biplot_title", "PCA Biplot")
    point_size  = float(params.get("point_size", 6))
    point_shape = params.get("point_shape", "circle")
    marker      = SHAPE_MAP.get(point_shape, "o")

    pc_x_lab = f"PC{pc_x_idx+1} ({explained[pc_x_idx]*100:.1f}%)"
    pc_y_lab = f"PC{pc_y_idx+1} ({explained[pc_y_idx]*100:.1f}%)"

    # Scale loadings to fit within score plot range
    scale_factor = (np.abs(scores[:, [pc_x_idx, pc_y_idx]]).max() /
                    np.abs(loadings[:, [pc_x_idx, pc_y_idx]]).max() * 0.65)

    fig, ax = plt.subplots(figsize=(7.8, 6.2), dpi=150)
    fig.patch.set_facecolor("#FAFAFA")
    _base_ax(ax, show_grid)

    # Scores
    if groups:
        unique_groups = list(dict.fromkeys(groups))
        color_map     = {g: PALETTE[i % len(PALETTE)] for i, g in enumerate(unique_groups)}
        for g in unique_groups:
            idx = [i for i, grp in enumerate(groups) if grp == g]
            ax.scatter(scores[idx, pc_x_idx], scores[idx, pc_y_idx],
                       c=color_map[g], s=point_size**2, marker=marker,
                       alpha=0.72, edgecolors="white", linewidths=0.5,
                       label=str(g), zorder=3)
        ax.legend(frameon=True, framealpha=0.88, fontsize=8,
                  edgecolor="#DDDDDD", loc="upper left")
    else:
        ax.scatter(scores[:, pc_x_idx], scores[:, pc_y_idx],
                   c=PALETTE[0], s=point_size**2, marker=marker,
                   alpha=0.72, edgecolors="white", linewidths=0.5, zorder=3)

    # Loadings (arrows)
    arrow_color = PALETTE[1]
    for i, var in enumerate(sel_cols):
        lx = loadings[i, pc_x_idx] * scale_factor
        ly = loadings[i, pc_y_idx] * scale_factor
        ax.annotate("", xy=(lx, ly), xytext=(0, 0),
                    arrowprops=dict(arrowstyle="-|>", color=arrow_color,
                                   lw=1.5, mutation_scale=10),
                    zorder=4)
        ax.text(lx * 1.08, ly * 1.08, var,
                fontsize=8, color=arrow_color, ha="center", va="center",
                fontweight=500, zorder=5)

    ax.axhline(0, color="#CCCCCC", linewidth=0.8, zorder=1)
    ax.axvline(0, color="#CCCCCC", linewidth=0.8, zorder=1)
    ax.set_xlabel(pc_x_lab, fontsize=10.5, labelpad=7, color="#333333")
    ax.set_ylabel(pc_y_lab, fontsize=10.5, labelpad=7, color="#333333")
    if title:
        ax.set_title(title, fontsize=11.5, fontweight="bold", color="#1C1C1C", pad=10)

    fig.tight_layout(pad=1.4)
    _save_fig(fig, session_dir, "biplot")


# ─── 4. Loading heatmap ───────────────────────────────────────────────────────

def _plot_loadings_heatmap(r: dict, params: dict, session_dir: Path) -> None:
    loadings  = np.array(r["loadings"])   # (p, k)
    sel_cols  = r["selected_cols"]
    pc_labels = r["pc_labels"]
    title     = params.get("loading_title", "PCA Loadings Heatmap")

    p, k = loadings.shape
    fig_h = max(4.5, p * 0.45 + 1.8)
    fig_w = max(5.5, k * 1.1 + 2.2)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h), dpi=150)
    fig.patch.set_facecolor("#FAFAFA")

    im = ax.imshow(loadings, aspect="auto", cmap="RdBu_r",
                   vmin=-1, vmax=1, interpolation="nearest")

    # Cell annotations
    for i in range(p):
        for j in range(k):
            val = loadings[i, j]
            ax.text(j, i, f"{val:.2f}",
                    ha="center", va="center", fontsize=8,
                    color="white" if abs(val) > 0.55 else "#444444")

    # Axes
    ax.set_xticks(range(k))
    ax.set_xticklabels(pc_labels, fontsize=9.5, color="#333333")
    ax.set_yticks(range(p))
    ax.set_yticklabels(sel_cols, fontsize=9, color="#333333")
    ax.set_xlabel("Principal Component", fontsize=10.5, labelpad=8, color="#333333")
    ax.set_ylabel("Variable", fontsize=10.5, labelpad=8, color="#333333")
    ax.tick_params(length=0)
    for spine in ax.spines.values():
        spine.set_visible(False)

    cbar = fig.colorbar(im, ax=ax, shrink=0.72, pad=0.02)
    cbar.set_label("Loading", fontsize=9, color="#555555")
    cbar.ax.tick_params(labelsize=8, colors="#999999")
    cbar.outline.set_edgecolor("#DDDDDD")

    if title:
        ax.set_title(title, fontsize=11.5, fontweight="bold", color="#1C1C1C", pad=10)

    fig.tight_layout(pad=1.6)
    _save_fig(fig, session_dir, "loading_heatmap")


# ─── Excel export ─────────────────────────────────────────────────────────────

def _save_excel(r: dict, session_dir: Path) -> None:
    scores    = np.array(r["scores"])
    loadings  = np.array(r["loadings"])
    explained = r["explained"]
    cumulative= r["cumulative"]
    pc_labels = r["pc_labels"]
    sel_cols  = r["selected_cols"]
    groups    = r.get("groups")

    wb   = openpyxl.Workbook()

    # ── Sheet 1: PC Scores ────────────────────────────────────────────────────
    ws1  = wb.active
    ws1.title = "PC Scores"

    header_fill = PatternFill("solid", fgColor="1C3557")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    sub_fill    = PatternFill("solid", fgColor="EAF3FB")
    sub_font    = Font(bold=True, color="1C3557", size=9)

    # Row 1: explained variance
    var_row = ([""] if not groups else ["", ""])  + [f"{v*100:.2f}%" for v in explained]
    ws1.append((["Sample"] if not groups else ["Sample", r["group_col"] or "Group"]) + pc_labels)
    # Style header
    for col_idx, cell in enumerate(ws1[1], start=1):
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row 2: variance %
    var_label_row = (["Var. explained"] if not groups else ["Var. explained", ""]) + \
                    [f"{v*100:.2f}%" for v in explained]
    ws1.append(var_label_row)
    for cell in ws1[2]:
        cell.font      = sub_font
        cell.fill      = sub_fill
        cell.alignment = Alignment(horizontal="center")

    # Row 3: cumulative %
    cum_label_row = (["Cumulative %"] if not groups else ["Cumulative %", ""]) + \
                    [f"{v*100:.2f}%" for v in cumulative]
    ws1.append(cum_label_row)
    for cell in ws1[3]:
        cell.font      = sub_font
        cell.fill      = sub_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for i, row_scores in enumerate(scores):
        row_data = [f"Sample_{i+1}"]
        if groups:
            row_data.append(groups[i])
        row_data += [round(float(v), 6) for v in row_scores]
        ws1.append(row_data)

    # Column widths
    ws1.column_dimensions["A"].width = 14
    if groups:
        ws1.column_dimensions["B"].width = 16
    for ci in range(len(pc_labels)):
        col_letter = openpyxl.utils.get_column_letter(ci + (3 if groups else 2))
        ws1.column_dimensions[col_letter].width = 13

    # ── Sheet 2: Loadings ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Loadings")
    ws2.append(["Variable"] + pc_labels)
    for cell in ws2[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Variance row
    ws2.append(["Var. explained"] + [f"{v*100:.2f}%" for v in explained])
    for cell in ws2[2]:
        cell.font      = sub_font
        cell.fill      = sub_fill
        cell.alignment = Alignment(horizontal="center")

    for i, var in enumerate(sel_cols):
        row_data = [var] + [round(float(loadings[i, j]), 6) for j in range(loadings.shape[1])]
        ws2.append(row_data)

    ws2.column_dimensions["A"].width = 22
    for ci in range(len(pc_labels)):
        col_letter = openpyxl.utils.get_column_letter(ci + 2)
        ws2.column_dimensions[col_letter].width = 13

    wb.save(session_dir / "pca_results.xlsx")
