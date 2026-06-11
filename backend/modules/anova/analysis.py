"""
ANOVA Suite — Analysis Engine v2
Designs:  CRD (1/2/3-way), RBD (1/2/3-way)
Post-hoc: Tukey HSD, Duncan, LSD
Plots:    Bar+error+CLD, Interaction plots
Export:   PNG + SVG + Excel (3 sheets)
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from itertools import combinations
from pathlib import Path
from scipy import stats
import statsmodels.api as sm
from statsmodels.formula.api import ols
from statsmodels.stats.anova import anova_lm
from statsmodels.stats.multicomp import pairwise_tukeyhsd
import scikit_posthocs as sp
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import re

PALETTE = [
    "#0072B2","#D55E00","#009E73","#E69F00",
    "#CC79A7","#56B4E9","#737373","#1C1C1C",
]

# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════

def _sanitize(name: str) -> str:
    s = re.sub(r"[^A-Za-z0-9_]", "_", str(name))
    if not s or s[0].isdigit():
        s = "X" + s
    if len(s) <= 2:
        s = "F_" + s
    return s


def build_formula(y_col, factor_cols, block_col, design):
    """
    Build patsy formula string.
    design: 'crd' | 'rbd'
    """
    sy   = _sanitize(y_col)
    sfac = [_sanitize(f) for f in factor_cols]
    sblk = _sanitize(block_col) if block_col else None

    terms = []
    # Block term (RBD only, additive)
    if design == "rbd" and sblk:
        terms.append(f"C({sblk})")

    # Main effects
    for sf in sfac:
        terms.append(f"C({sf})")

    # Interactions between treatment factors only
    if len(sfac) >= 2:
        for i, j in combinations(range(len(sfac)), 2):
            terms.append(f"C({sfac[i]}):C({sfac[j]})")
    if len(sfac) == 3:
        terms.append(f"C({sfac[0]}):C({sfac[1]}):C({sfac[2]})")

    formula = sy + " ~ " + " + ".join(terms)
    return formula, sy, sfac, sblk


def model_formula_display(y_col, factor_cols, block_col, design):
    """
    Human-readable formula for display in the UI.
    """
    terms = []
    if design == "rbd" and block_col:
        terms.append(f"C({block_col})")
    for f in factor_cols:
        terms.append(f"C({f})")
    if len(factor_cols) >= 2:
        for i, j in combinations(range(len(factor_cols)), 2):
            terms.append(f"C({factor_cols[i]}):C({factor_cols[j]})")
    if len(factor_cols) == 3:
        terms.append(f"C({factor_cols[0]}):C({factor_cols[1]}):C({factor_cols[2]})")
    return y_col + " ~ " + " + ".join(terms)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN API
# ══════════════════════════════════════════════════════════════════════════════

def run_anova(data, y_col, factor_cols, block_col,
              design, posthoc, error_bar, params, session_dir):
    """
    design:    'crd' | 'rbd'
    posthoc:   'tukey' | 'duncan' | 'lsd'
    error_bar: 'sem'   | 'sd'     | 'ci95'
    """
    df = pd.DataFrame(data)
    df[y_col] = pd.to_numeric(df[y_col], errors="coerce")
    for f in factor_cols:
        df[f] = df[f].astype(str)
    if block_col:
        df[block_col] = df[block_col].astype(str)

    used = [y_col] + factor_cols + ([block_col] if block_col else [])
    df   = df[used].dropna()
    n    = len(df)
    if n < 4:
        raise ValueError(f"Need at least 4 complete rows (got {n})")

    # Build rename map and work on renamed copy
    rename = {y_col: _sanitize(y_col)}
    rename.update({f: _sanitize(f) for f in factor_cols})
    if block_col:
        rename[block_col] = _sanitize(block_col)
    df2 = df.rename(columns=rename)

    formula, sy, sfac, sblk = build_formula(y_col, factor_cols, block_col, design)

    model    = ols(formula, data=df2).fit()
    anova_df = anova_lm(model, typ=2)

    # ── Build ANOVA table ─────────────────────────────────────────────────────
    anova_table = []
    for src, row in anova_df.iterrows():
        # Restore human-readable label
        label = src
        for sf, f in rename.items():
            label = label.replace(f"C({f})", f"C({sf})")
        label = label.replace(":", " × ").strip()

        f_val = float(row["F"])       if not pd.isna(row["F"])        else None
        p_val = float(row["PR(>F)"]) if not pd.isna(row["PR(>F)"])   else None
        df_v  = float(row["df"])
        ss_v  = float(row["sum_sq"])

        anova_table.append({
            "source": label,
            "ss":  ss_v,
            "df":  df_v,
            "ms":  ss_v / df_v if df_v != 0 else None,
            "f":   f_val,
            "p":   p_val,
        })

    # Residual row
    if not any(r["source"] == "Residual" for r in anova_table):
        ss_r = float(model.ssr)
        df_r = float(model.df_resid)
        anova_table.append({
            "source": "Residual",
            "ss": ss_r, "df": df_r,
            "ms": ss_r / df_r if df_r > 0 else None,
            "f": None, "p": None,
        })

    # ── Post-hoc on primary treatment factor ──────────────────────────────────
    ph_results, cld = _posthoc(df, y_col, factor_cols[0], posthoc)

    # ── Group means ───────────────────────────────────────────────────────────
    means_table = _group_means(df, y_col, factor_cols, error_bar)

    # ── Model formula for display ─────────────────────────────────────────────
    formula_display = model_formula_display(y_col, factor_cols, block_col, design)

    result = {
        "y_col":           y_col,
        "factor_cols":     factor_cols,
        "block_col":       block_col,
        "design":          design,
        "posthoc":         posthoc,
        "error_bar":       error_bar,
        "n":               int(n),
        "formula_display": formula_display,
        "anova_table":     anova_table,
        "ph_results":      ph_results,
        "cld":             cld,
        "means_table":     means_table,
        "raw_data":        df.to_dict(orient="list"),
    }

    _draw_all(result, params, session_dir)
    _save_excel(result, session_dir)
    return result


def regenerate_plots(stored, params, session_dir):
    _draw_all(stored, params, session_dir)


# ══════════════════════════════════════════════════════════════════════════════
# POST-HOC
# ══════════════════════════════════════════════════════════════════════════════

def _posthoc(df, y_col, factor, method):
    groups     = sorted(df[factor].unique(), key=str)
    if len(groups) < 2:
        return [], {g: "a" for g in groups}

    if method == "tukey":
        res = pairwise_tukeyhsd(df[y_col], df[factor])
        ph  = []
        for row in res.summary().data[1:]:
            ph.append({
                "group1":    str(row[0]),
                "group2":    str(row[1]),
                "mean_diff": float(row[2]),
                "p_adj":     float(row[3]),
                "lower":     float(row[4]),
                "upper":     float(row[5]),
                "reject":    bool(row[6]),
            })
    elif method == "lsd":
        ph = _lsd_test(df, y_col, factor)
    else:
        ph = _duncan_test(df, y_col, factor)

    cld = _cld(groups, ph)
    return ph, cld


def _lsd_test(df, y_col, factor):
    groups      = sorted(df[factor].unique(), key=str)
    n_total     = len(df)
    n_groups    = len(groups)
    data_by_grp = {g: df.loc[df[factor] == g, y_col].dropna().values for g in groups}

    ss_within = sum(np.sum((v - v.mean())**2) for v in data_by_grp.values())
    df_within = n_total - n_groups
    mse       = ss_within / df_within if df_within > 0 else np.nan

    ph = []
    for g1, g2 in combinations(groups, 2):
        v1, v2  = data_by_grp[g1], data_by_grp[g2]
        n1, n2  = len(v1), len(v2)
        diff    = float(v1.mean() - v2.mean())
        se      = np.sqrt(mse * (1/n1 + 1/n2)) if not np.isnan(mse) else np.nan
        t_stat  = diff / se if (not np.isnan(se) and se > 0) else np.nan
        p_val   = float(2 * stats.t.sf(abs(t_stat), df=df_within)) if not np.isnan(t_stat) else 1.0
        t_crit  = stats.t.ppf(0.975, df=df_within)
        lsd_val = t_crit * se if not np.isnan(se) else np.nan
        ph.append({
            "group1":    str(g1), "group2": str(g2),
            "mean_diff": diff,
            "p_adj":     p_val,
            "lower":     diff - lsd_val if not np.isnan(lsd_val) else None,
            "upper":     diff + lsd_val if not np.isnan(lsd_val) else None,
            "reject":    p_val < 0.05,
        })
    return ph


def _duncan_test(df, y_col, factor):
    try:
        ph_mat = sp.posthoc_duncan(df, val_col=y_col, group_col=factor)
    except Exception:
        ph_mat = sp.posthoc_dunn(df, val_col=y_col, group_col=factor,
                                 p_adjust="bonferroni")
    ph = []
    grp_list = list(ph_mat.columns)
    data_by_grp = {g: df.loc[df[factor] == g, y_col].dropna().values
                   for g in grp_list}
    for g1, g2 in combinations(grp_list, 2):
        p_val = float(ph_mat.loc[g1, g2])
        v1    = data_by_grp.get(g1, np.array([]))
        v2    = data_by_grp.get(g2, np.array([]))
        diff  = float(v1.mean() - v2.mean()) if len(v1) > 0 and len(v2) > 0 else 0.0
        ph.append({
            "group1": str(g1), "group2": str(g2),
            "mean_diff": diff,
            "p_adj": p_val,
            "lower": None, "upper": None,
            "reject": p_val < 0.05,
        })
    return ph


def _cld(groups, ph_results):
    """Compact Letter Display — groups sharing a letter are not significantly different."""
    groups   = [str(g) for g in groups]
    ns_pairs = set()
    for row in ph_results:
        if not row["reject"]:
            ns_pairs.add((str(row["group1"]), str(row["group2"])))
            ns_pairs.add((str(row["group2"]), str(row["group1"])))

    letters  = "abcdefghijklmnopqrstuvwxyz"
    assigned = {g: set() for g in groups}
    letter_i = 0

    for g in groups:
        placed = False
        for li in range(letter_i):
            ltr     = letters[li]
            members = [gg for gg in groups if ltr in assigned[gg]]
            if all((g, m) in ns_pairs for m in members):
                assigned[g].add(ltr)
                placed = True
                break
        if not placed:
            assigned[g].add(letters[letter_i])
            letter_i += 1

    return {g: "".join(sorted(assigned[g])) for g in groups}


# ══════════════════════════════════════════════════════════════════════════════
# GROUP MEANS
# ══════════════════════════════════════════════════════════════════════════════

def _group_means(df, y_col, factor_cols, error_bar):
    rows = []
    for keys, grp in df.groupby(factor_cols):
        vals = grp[y_col].dropna().values
        if len(vals) == 0:
            continue
        n    = len(vals)
        mean = float(np.mean(vals))
        sd   = float(np.std(vals, ddof=1)) if n > 1 else 0.0
        sem  = sd / np.sqrt(n)
        ci95 = 1.96 * sem
        err  = sd if error_bar == "sd" else (ci95 if error_bar == "ci95" else sem)

        key_list = [keys] if isinstance(keys, str) else list(keys)
        row = {f: k for f, k in zip(factor_cols, key_list)}
        row.update({"mean": mean, "sd": sd, "sem": sem,
                    "ci95": ci95, "err": err, "n": n})
        rows.append(row)
    return rows


# ══════════════════════════════════════════════════════════════════════════════
# PLOTTING
# ══════════════════════════════════════════════════════════════════════════════

def _ax_style(ax, grid=True):
    ax.set_facecolor("#FAFAFA")
    ax.spines[["top","right"]].set_visible(False)
    ax.spines["left"].set_color("#CCCCCC")
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.tick_params(colors="#999999", labelsize=9)
    ax.grid(grid, axis="y", linestyle="--", linewidth=0.45,
            color="#DDDDDD", alpha=0.8, zorder=0)


def _save_fig(fig, session_dir, name):
    fig.savefig(session_dir / f"{name}.png", dpi=150,
                bbox_inches="tight", facecolor=fig.get_facecolor())
    fig.savefig(session_dir / f"{name}.svg", format="svg",
                bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)


def _draw_all(r, params, session_dir):
    _plot_bar(r, params, session_dir)
    if len(r["factor_cols"]) >= 2:
        _plot_interaction(r, params, session_dir)


# ─── Bar + error bars + CLD ───────────────────────────────────────────────────

def _plot_bar(r, params, session_dir):
    means       = r["means_table"]
    cld         = r["cld"]
    factor_cols = r["factor_cols"]
    y_col       = r["y_col"]
    error_bar   = r["error_bar"]
    show_grid   = bool(params.get("show_grid", True))
    title       = params.get("bar_title", f"Mean {y_col}")
    font_size   = float(params.get("font_size", 10))
    eb_label    = {"sem":"SEM","sd":"SD","ci95":"95% CI"}[error_bar]
    user_colors = params.get("bar_colors") or PALETTE
    if not user_colors:
        user_colors = PALETTE

    df_m = pd.DataFrame(means)
    f0   = factor_cols[0]

    if len(factor_cols) == 1:
        groups    = sorted(df_m[f0].unique(), key=str)
        n_grps    = len(groups)
        colors    = (user_colors * n_grps)[:n_grps]
        x         = np.arange(n_grps)

        fig, ax = plt.subplots(figsize=(max(6, n_grps * 0.95 + 2), 5.5), dpi=150)
        fig.patch.set_facecolor("#FAFAFA")
        _ax_style(ax, show_grid)

        max_top = 0
        for i, grp in enumerate(groups):
            row       = df_m[df_m[f0] == grp].iloc[0]
            mean, err = row["mean"], row["err"]
            ax.bar(x[i], mean, color=colors[i], alpha=0.84,
                   width=0.6, zorder=3, edgecolor="white", linewidth=0.8)
            ax.errorbar(x[i], mean, yerr=err, fmt="none",
                        color="#333", linewidth=1.6, capsize=5,
                        capthick=1.6, zorder=4)
            max_top = max(max_top, mean + err)

        letter_y = max_top * 1.06
        for i, grp in enumerate(groups):
            ax.text(x[i], letter_y, cld.get(str(grp), ""),
                    ha="center", va="bottom", fontsize=font_size,
                    fontweight="bold", color="#333", zorder=5)

        ax.set_xticks(x)
        ax.set_xticklabels(groups, fontsize=font_size - 0.5,
                           rotation=20 if n_grps > 6 else 0)
        ax.set_ylim(0, letter_y * 1.14)

    else:
        # Multi-factor: f0 on x-axis, f1 as colour groups
        f1      = factor_cols[1]
        grps0   = sorted(df_m[f0].unique(), key=str)
        grps1   = sorted(df_m[f1].unique(), key=str)
        n0, n1  = len(grps0), len(grps1)
        width   = 0.8 / n1
        x       = np.arange(n0)
        colors  = (user_colors * n1)[:n1]

        fig, ax = plt.subplots(
            figsize=(max(7, n0 * n1 * 0.6 + 2), 5.5), dpi=150)
        fig.patch.set_facecolor("#FAFAFA")
        _ax_style(ax, show_grid)

        max_top = 0
        for j, g1 in enumerate(grps1):
            offset = (j - (n1 - 1) / 2) * width
            for i, g0 in enumerate(grps0):
                mask = (df_m[f0] == g0) & (df_m[f1] == g1)
                sub  = df_m[mask]
                if sub.empty:
                    continue
                row      = sub.iloc[0]
                mean, err = row["mean"], row["err"]
                ax.bar(x[i] + offset, mean, color=colors[j],
                       alpha=0.84, width=width * 0.88,
                       zorder=3, edgecolor="white", linewidth=0.6)
                ax.errorbar(x[i] + offset, mean, yerr=err, fmt="none",
                            color="#333", linewidth=1.2,
                            capsize=4, capthick=1.2, zorder=4)
                max_top = max(max_top, mean + err)

        # CLD for f0 above each group's tallest bar
        for i, g0 in enumerate(grps0):
            ltr = cld.get(str(g0), "")
            sub = df_m[df_m[f0] == g0]
            top = (sub["mean"] + sub["err"]).max() if not sub.empty else max_top
            ax.text(x[i], top * 1.04, ltr,
                    ha="center", va="bottom", fontsize=font_size,
                    fontweight="bold", color="#333", zorder=5)

        ax.set_xticks(x)
        ax.set_xticklabels(grps0, fontsize=font_size - 0.5,
                           rotation=20 if n0 > 5 else 0)
        ax.set_ylim(0, max_top * 1.22)

        patches = [mpatches.Patch(color=colors[j], label=str(g1))
                   for j, g1 in enumerate(grps1)]
        ax.legend(handles=patches, title=f1, fontsize=font_size - 1,
                  frameon=True, framealpha=0.88, edgecolor="#DDDDDD")

    design_label = "RBD" if r["design"] == "rbd" else "CRD"
    ax.set_xlabel(params.get("x_label", f0),
                  fontsize=font_size + 0.5, labelpad=7, color="#333")
    ax.set_ylabel(params.get("y_label", f"Mean {y_col} ± {eb_label}"),
                  fontsize=font_size + 0.5, labelpad=7, color="#333")
    if title:
        ax.set_title(title, fontsize=font_size + 2,
                     fontweight="bold", color="#1C1C1C", pad=10)

    ax.text(0.01, -0.13,
            f"Design: {design_label}  ·  Error bars = {eb_label}  ·  "
            f"Letters = post-hoc groups (p<0.05)",
            transform=ax.transAxes, fontsize=font_size - 2,
            color="#AAAAAA", va="top")

    fig.tight_layout(pad=1.6)
    _save_fig(fig, session_dir, "bar_plot")


# ─── Interaction plot ─────────────────────────────────────────────────────────

def _plot_interaction(r, params, session_dir):
    df        = pd.DataFrame(r["raw_data"])
    y_col     = r["y_col"]
    facs      = r["factor_cols"]
    show_grid = bool(params.get("show_grid", True))
    font_size = float(params.get("font_size", 10))
    title     = params.get("int_title", f"Interaction Plot — {y_col}")
    error_bar = r["error_bar"]
    eb_label  = {"sem":"SEM","sd":"SD","ci95":"95% CI"}[error_bar]
    user_colors = params.get("bar_colors") or PALETTE

    if len(facs) == 2:
        _int_2way(df, y_col, facs, title, font_size, show_grid,
                  error_bar, eb_label, user_colors, session_dir)
    else:
        _int_3way(df, y_col, facs, title, font_size, show_grid,
                  error_bar, eb_label, user_colors, session_dir)


def _int_2way(df, y_col, facs, title, font_size, show_grid,
               error_bar, eb_label, user_colors, session_dir):
    f0, f1  = facs[0], facs[1]
    grps0   = sorted(df[f0].unique(), key=str)
    grps1   = sorted(df[f1].unique(), key=str)
    x       = np.arange(len(grps0))

    fig, ax = plt.subplots(figsize=(7.5, 5.5), dpi=150)
    fig.patch.set_facecolor("#FAFAFA")
    _ax_style(ax, show_grid)

    for j, g1 in enumerate(grps1):
        col  = user_colors[j % len(user_colors)]
        ys, errs = [], []
        for g0 in grps0:
            vals = df[(df[f0]==g0) & (df[f1]==g1)][y_col].dropna().values
            if len(vals) == 0:
                ys.append(np.nan); errs.append(0); continue
            m  = np.mean(vals)
            sd = np.std(vals, ddof=1) if len(vals) > 1 else 0
            se = sd / np.sqrt(len(vals))
            ys.append(m)
            errs.append(sd if error_bar=="sd" else
                        (1.96*se if error_bar=="ci95" else se))
        ax.plot(x, ys, color=col, linewidth=2, marker="o",
                markersize=7, label=str(g1), zorder=3)
        ax.errorbar(x, ys, yerr=errs, fmt="none",
                    color=col, linewidth=1.2, capsize=4,
                    capthick=1.2, alpha=0.65, zorder=2)

    ax.set_xticks(x)
    ax.set_xticklabels(grps0, fontsize=font_size - 0.5)
    ax.set_xlabel(f0, fontsize=font_size + 0.5, labelpad=7, color="#333")
    ax.set_ylabel(f"Mean {y_col} ± {eb_label}",
                  fontsize=font_size + 0.5, labelpad=7, color="#333")
    ax.legend(title=f1, fontsize=font_size-1,
              frameon=True, framealpha=0.88, edgecolor="#DDDDDD")
    if title:
        ax.set_title(title, fontsize=font_size+2,
                     fontweight="bold", color="#1C1C1C", pad=10)
    fig.tight_layout(pad=1.6)
    _save_fig(fig, session_dir, "interaction_plot")


def _int_3way(df, y_col, facs, title, font_size, show_grid,
               error_bar, eb_label, user_colors, session_dir):
    f0, f1, f2 = facs[0], facs[1], facs[2]
    levels2    = sorted(df[f2].unique(), key=str)
    n_sub      = len(levels2)
    grps0      = sorted(df[f0].unique(), key=str)
    grps1      = sorted(df[f1].unique(), key=str)
    x          = np.arange(len(grps0))

    fig, axes = plt.subplots(1, n_sub,
                             figsize=(6*n_sub, 5.5), dpi=150, sharey=True)
    if n_sub == 1:
        axes = [axes]
    fig.patch.set_facecolor("#FAFAFA")

    for si, lv2 in enumerate(levels2):
        ax  = axes[si]
        sub = df[df[f2] == lv2]
        _ax_style(ax, show_grid)
        for j, g1 in enumerate(grps1):
            col = user_colors[j % len(user_colors)]
            ys, errs = [], []
            for g0 in grps0:
                vals = sub[(sub[f0]==g0) & (sub[f1]==g1)][y_col].dropna().values
                if len(vals) == 0:
                    ys.append(np.nan); errs.append(0); continue
                m  = np.mean(vals)
                sd = np.std(vals, ddof=1) if len(vals) > 1 else 0
                se = sd / np.sqrt(len(vals))
                ys.append(m)
                errs.append(sd if error_bar=="sd" else
                            (1.96*se if error_bar=="ci95" else se))
            ax.plot(x, ys, color=col, linewidth=2, marker="o",
                    markersize=6, label=str(g1), zorder=3)
            ax.errorbar(x, ys, yerr=errs, fmt="none",
                        color=col, linewidth=1.2, capsize=4,
                        capthick=1.2, alpha=0.65, zorder=2)
        ax.set_xticks(x)
        ax.set_xticklabels(grps0, fontsize=font_size-1)
        ax.set_title(f"{f2} = {lv2}", fontsize=font_size, color="#555")
        ax.set_xlabel(f0, fontsize=font_size, color="#333")
        if si == 0:
            ax.set_ylabel(f"Mean {y_col} ± {eb_label}",
                          fontsize=font_size, color="#333")
        ax.legend(title=f1, fontsize=font_size-2,
                  frameon=True, framealpha=0.88, edgecolor="#DDDDDD")

    if title:
        fig.suptitle(title, fontsize=font_size+3,
                     fontweight="bold", color="#1C1C1C", y=1.01)
    fig.tight_layout(pad=1.6)
    _save_fig(fig, session_dir, "interaction_plot")


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def _save_excel(r, session_dir):
    wb    = openpyxl.Workbook()
    hf    = PatternFill("solid", fgColor="1C3557")
    hfont = Font(bold=True, color="FFFFFF", size=10)
    ctr   = Alignment(horizontal="center")

    def hr(ws, vals):
        ws.append(vals)
        for c in ws[ws.max_row]:
            c.font = hfont; c.fill = hf; c.alignment = ctr

    def sig(p):
        if p is None: return ""
        return "***" if p<0.001 else "**" if p<0.01 else "*" if p<0.05 else "ns"

    # Sheet 1 — ANOVA table
    ws1 = wb.active; ws1.title = "ANOVA Table"
    hr(ws1, ["Source","SS","df","MS","F","p-value","Sig"])
    for row in r["anova_table"]:
        ws1.append([
            row["source"],
            round(row["ss"],6) if row["ss"] is not None else "—",
            round(row["df"],1) if row["df"] is not None else "—",
            round(row["ms"],6) if row["ms"] is not None else "—",
            round(row["f"], 4) if row["f"]  is not None else "—",
            round(row["p"], 6) if row["p"]  is not None else "—",
            sig(row["p"]),
        ])
    # Model formula note
    ws1.append([])
    ws1.append(["Model:", r.get("formula_display","")])
    ws1.append(["Design:", r["design"].upper()])
    for col, w in zip("ABCDEFG",[30,14,8,14,12,12,8]):
        ws1.column_dimensions[col].width = w

    # Sheet 2 — Means & CLD
    ws2 = wb.create_sheet("Means & CLD")
    facs    = r["factor_cols"]
    headers = facs + ["n","Mean","SD","SEM","95% CI","CLD"]
    hr(ws2, headers)
    for row in r["means_table"]:
        cld_str = r["cld"].get(str(row.get(facs[0],"")), "") \
                  if len(facs) == 1 else ""
        ws2.append(
            [row.get(f,"") for f in facs] +
            [row["n"], round(row["mean"],4), round(row["sd"],4),
             round(row["sem"],4), round(row["ci95"],4), cld_str]
        )
    for i, w in enumerate([20]*len(facs)+[6,12,12,12,12,8]):
        ws2.column_dimensions[
            openpyxl.utils.get_column_letter(i+1)].width = w

    # Sheet 3 — Post-hoc
    ws3 = wb.create_sheet("Post-hoc")
    hr(ws3, ["Group 1","Group 2","Mean Diff","p-adj","Sig","Reject H0"])
    for row in r["ph_results"]:
        ws3.append([
            row["group1"], row["group2"],
            round(row["mean_diff"],4),
            round(row["p_adj"],6),
            sig(row["p_adj"]),
            "Yes" if row["reject"] else "No",
        ])
    for col, w in zip("ABCDEF",[16,16,14,14,8,10]):
        ws3.column_dimensions[col].width = w

    wb.save(session_dir / "anova_results.xlsx")
